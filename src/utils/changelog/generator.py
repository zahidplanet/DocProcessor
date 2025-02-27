"""
Module for generating changelogs between document versions.
"""

import difflib
import os
from datetime import datetime
from typing import List, Dict, Any, Optional, Tuple

import pandas as pd
from docx import Document as DocxDocument
from PyPDF2 import PdfReader
import docx2txt
from openpyxl import load_workbook

from src.models.document import Document, Comment


class ChangelogGenerator:
    """
    Class for generating changelogs between document versions.
    Supports multiple document formats including PDF, DOCX, DOC, XLSX, and XLS.
    """
    
    def __init__(self, uploads_dir: str):
        """
        Initialize the changelog generator.
        
        Args:
            uploads_dir: Directory where uploaded documents are stored
        """
        self.uploads_dir = uploads_dir
    
    def generate_changelog(self, original_doc: Document, new_doc: Document, 
                           resolved_comments: Optional[List[Comment]] = None) -> Dict[str, Any]:
        """
        Generate a changelog between two document versions.
        
        Args:
            original_doc: The original document
            new_doc: The new document version
            resolved_comments: List of resolved comments between versions
            
        Returns:
            Dictionary containing change information
        """
        file_ext = original_doc.filename.split('.')[-1].lower()
        
        # Get content from both documents
        original_content = self._extract_content(original_doc.filename, file_ext)
        new_content = self._extract_content(new_doc.filename, file_ext)
        
        # Calculate differences
        changes = self._calculate_differences(original_content, new_content, file_ext)
        
        # Prepare changelog
        changelog = {
            "original_document": {
                "id": original_doc.id,
                "filename": original_doc.original_filename,
                "version": original_doc.version
            },
            "new_document": {
                "id": new_doc.id,
                "filename": new_doc.original_filename,
                "version": new_doc.version
            },
            "changes": changes,
            "timestamp": datetime.now().isoformat(),
            "resolved_comments": []
        }
        
        # Add resolved comments if provided
        if resolved_comments:
            changelog["resolved_comments"] = [
                {
                    "id": comment.id,
                    "text": comment.text,
                    "resolution": comment.resolution_text,
                    "resolved_by": comment.resolved_by,
                    "resolved_at": comment.resolved_at.isoformat() if comment.resolved_at else None
                }
                for comment in resolved_comments
            ]
        
        return changelog
    
    def generate_formatted_changelog(self, original_doc: Document, new_doc: Document,
                                    resolved_comments: Optional[List[Comment]] = None) -> str:
        """
        Generate a formatted text changelog between two document versions.
        
        Args:
            original_doc: The original document
            new_doc: The new document version
            resolved_comments: List of resolved comments between versions
            
        Returns:
            Formatted changelog as a string
        """
        changelog = self.generate_changelog(original_doc, new_doc, resolved_comments)
        
        # Format the changelog as text
        formatted = f"# Changelog: {original_doc.original_filename} (v{original_doc.version} → v{new_doc.version})\n\n"
        formatted += f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n\n"
        
        formatted += "## Document Information\n"
        formatted += f"- Original: {original_doc.original_filename} (v{original_doc.version})\n"
        formatted += f"- New: {new_doc.original_filename} (v{new_doc.version})\n\n"
        
        formatted += "## Changes\n"
        
        if isinstance(changelog["changes"], dict) and "summary" in changelog["changes"]:
            # For formats where we provide a summary
            formatted += changelog["changes"]["summary"] + "\n\n"
        else:
            # For text-based formats with line-by-line changes
            for change in changelog["changes"]:
                if change["type"] == "addition":
                    formatted += f"+ Added: {change['content']}\n"
                elif change["type"] == "deletion":
                    formatted += f"- Removed: {change['content']}\n"
                elif change["type"] == "modification":
                    formatted += f"* Modified: {change['old']} → {change['new']}\n"
        
        if resolved_comments:
            formatted += "\n## Resolved Comments\n"
            for comment in changelog["resolved_comments"]:
                formatted += f"- Comment: {comment['text']}\n"
                formatted += f"  Resolution: {comment['resolution']}\n"
                formatted += f"  Resolved by: {comment['resolved_by']}\n\n"
        
        return formatted
    
    def _extract_content(self, filename: str, file_ext: str) -> Any:
        """
        Extract content from a document based on its file extension.
        
        Args:
            filename: Name of the file
            file_ext: File extension
            
        Returns:
            Extracted content in appropriate format
        """
        filepath = os.path.join(self.uploads_dir, filename)
        
        if file_ext in ['pdf']:
            return self._extract_pdf_content(filepath)
        elif file_ext in ['docx', 'doc']:
            return self._extract_word_content(filepath)
        elif file_ext in ['xlsx', 'xls']:
            return self._extract_excel_content(filepath)
        else:
            # Default to treating as text
            with open(filepath, 'r', errors='ignore') as f:
                return f.read()
    
    def _extract_pdf_content(self, filepath: str) -> str:
        """Extract text content from a PDF file."""
        reader = PdfReader(filepath)
        text = ""
        for page in reader.pages:
            text += page.extract_text() + "\n"
        return text
    
    def _extract_word_content(self, filepath: str) -> str:
        """Extract text content from a Word document."""
        try:
            # Try docx2txt first (handles both .doc and .docx)
            return docx2txt.process(filepath)
        except:
            # Fallback to python-docx (only handles .docx)
            try:
                doc = DocxDocument(filepath)
                return "\n".join([para.text for para in doc.paragraphs])
            except:
                return "Error extracting Word document content"
    
    def _extract_excel_content(self, filepath: str) -> pd.DataFrame:
        """Extract content from an Excel file."""
        try:
            # For .xlsx files
            workbook = load_workbook(filepath)
            
            # Process all sheets
            sheets_data = {}
            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                data = []
                for row in sheet.iter_rows(values_only=True):
                    data.append(row)
                sheets_data[sheet_name] = data
            
            return sheets_data
        except:
            # Fallback to pandas
            try:
                return pd.read_excel(filepath, sheet_name=None)
            except:
                return {"Error": "Could not extract Excel content"}
    
    def _calculate_differences(self, original_content: Any, new_content: Any, file_ext: str) -> List[Dict[str, Any]]:
        """
        Calculate differences between document contents.
        
        Args:
            original_content: Content from the original document
            new_content: Content from the new document
            file_ext: File extension to determine processing method
            
        Returns:
            List of changes or a summary dictionary
        """
        if file_ext in ['pdf', 'docx', 'doc']:
            return self._text_differences(original_content, new_content)
        elif file_ext in ['xlsx', 'xls']:
            return self._excel_differences(original_content, new_content)
        else:
            # Default to text comparison
            return self._text_differences(original_content, new_content)
    
    def _text_differences(self, original_text: str, new_text: str) -> List[Dict[str, Any]]:
        """Calculate differences between text documents."""
        # Split into lines
        original_lines = original_text.splitlines()
        new_lines = new_text.splitlines()
        
        # Generate diff
        diff = list(difflib.unified_diff(original_lines, new_lines, lineterm=''))
        
        # Parse the diff output
        changes = []
        i = 0
        while i < len(diff):
            line = diff[i]
            if line.startswith('+') and not line.startswith('+++'):
                changes.append({
                    "type": "addition",
                    "content": line[1:].strip()
                })
            elif line.startswith('-') and not line.startswith('---'):
                # Check if this is part of a modification
                if (i+1 < len(diff) and diff[i+1].startswith('+') and 
                    not diff[i+1].startswith('+++')):
                    changes.append({
                        "type": "modification",
                        "old": line[1:].strip(),
                        "new": diff[i+1][1:].strip()
                    })
                    i += 1  # Skip the next line as we've already processed it
                else:
                    changes.append({
                        "type": "deletion",
                        "content": line[1:].strip()
                    })
            i += 1
        
        return changes
    
    def _excel_differences(self, original_data: Dict[str, Any], new_data: Dict[str, Any]) -> Dict[str, Any]:
        """Calculate differences between Excel documents."""
        changes = {
            "summary": "Changes detected in Excel document",
            "details": {},
            "sheet_changes": []
        }
        
        # Check for added or removed sheets
        original_sheets = set(original_data.keys())
        new_sheets = set(new_data.keys())
        
        added_sheets = new_sheets - original_sheets
        removed_sheets = original_sheets - new_sheets
        common_sheets = original_sheets.intersection(new_sheets)
        
        if added_sheets:
            changes["sheet_changes"].append({
                "type": "addition",
                "sheets": list(added_sheets)
            })
        
        if removed_sheets:
            changes["sheet_changes"].append({
                "type": "deletion",
                "sheets": list(removed_sheets)
            })
        
        # Compare content of common sheets
        for sheet in common_sheets:
            original_sheet = original_data[sheet]
            new_sheet = new_data[sheet]
            
            # Create a summary of cell changes
            cell_changes = []
            
            # Convert to a format we can compare
            if isinstance(original_sheet, list) and isinstance(new_sheet, list):
                # Handle list of tuples format
                for i, (orig_row, new_row) in enumerate(zip(original_sheet, new_sheet)):
                    for j, (orig_cell, new_cell) in enumerate(zip(orig_row, new_row)):
                        if orig_cell != new_cell:
                            cell_changes.append({
                                "row": i+1,
                                "column": j+1,
                                "old_value": str(orig_cell) if orig_cell is not None else "",
                                "new_value": str(new_cell) if new_cell is not None else ""
                            })
            elif isinstance(original_sheet, pd.DataFrame) and isinstance(new_sheet, pd.DataFrame):
                # Handle pandas DataFrame format
                # Find differences
                if not original_sheet.equals(new_sheet):
                    # Check shape differences
                    if original_sheet.shape != new_sheet.shape:
                        changes["details"][sheet] = f"Sheet shape changed: {original_sheet.shape} -> {new_sheet.shape}"
                    
                    # Check data differences
                    for col in set(original_sheet.columns).intersection(set(new_sheet.columns)):
                        for idx in set(original_sheet.index).intersection(set(new_sheet.index)):
                            if original_sheet.loc[idx, col] != new_sheet.loc[idx, col]:
                                cell_changes.append({
                                    "row": idx+1 if isinstance(idx, int) else str(idx),
                                    "column": col,
                                    "old_value": str(original_sheet.loc[idx, col]),
                                    "new_value": str(new_sheet.loc[idx, col])
                                })
            
            if cell_changes:
                changes["details"][sheet] = {
                    "cell_changes": cell_changes
                }
        
        return changes 